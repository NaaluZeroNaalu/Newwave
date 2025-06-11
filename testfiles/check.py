import streamlit as st
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta

st.title("NCR Report Excel Generator")

# Step 1: Get JSON input from user
json_input = st.text_area("Paste your JSON data here (as a list of dictionaries):")

uploaded_file = st.file_uploader("Or upload a JSON file", type=["json"])

if json_input or uploaded_file:
    try:
        if uploaded_file:
            data = json.load(uploaded_file)
        else:
            data = json.loads(json_input)

        # Current date and time (04:39 PM IST on May 29, 2025)
        current_date = datetime(2025, 5, 29, 16, 39)

        # Define the 21-day threshold
        threshold_date = current_date - timedelta(days=21)

        # Map JSON disciplines to table disciplines
        discipline_mapping = {
            'FW': 'Finishing',  # Assuming FW maps to Finishing
            'SW': 'Civil',      # Assuming SW (Shear Wall) maps to Civil
            'HSE': 'Works'      # Assuming HSE maps to Works
        }

        # Initialize dictionaries to store counts for open and resolved NCRs
        sites = ['TOWER F', 'F1', 'F2', 'COMMON', 'TOWER G', 'G1', 'G2', 'G3', 'COMMON', 'TOWER H', 'H1', 'H2', 'H3', 'H4', 'COMMON']
        open_ncr_data = {site: {'Civil': 0, 'Finishing': 0, 'Works': 0, 'MEP': 0} for site in sites}
        resolved_ncr_data = {site: {'Civil': 0, 'Finishing': 0, 'Works': 0, 'MEP': 0} for site in sites}

        # Process JSON data to count open and resolved NCRs within 21 days
        for entry in data:
            created_date = datetime.strptime(entry['Created Date'], '%Y-%m-%d %H:%M:%S')
            if created_date >= threshold_date:
                tower = entry['Tower']
                discipline = discipline_mapping.get(entry['Discipline'], 'MEP')
                if tower == 'F':
                    site = 'TOWER F'
                elif tower == 'G':
                    site = 'TOWER G'
                elif tower == 'H':
                    site = 'TOWER H'
                else:
                    continue
                if entry['Status'] == 'Open':
                    open_ncr_data[site][discipline] += 1
                elif entry['Status'] == 'Closed':
                    resolved_ncr_data[site][discipline] += 1

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "NCR Report"

        # Define headers (corrected to match screenshot exactly)
        headers = [
            ["SITE", "NCR resolved beyond 21 days", "", "", "", "NCR open beyond 21 days", "", "", "", "TOTAL"],
            ["", "Civil", "Finishing", "Works", "MEP", "Civil", "Finishing", "Works", "MEP", ""]
        ]

        # Write headers
        for row_idx, row in enumerate(headers, 1):
            for col_idx, cell in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell)

        # Merge cells as per the screenshot
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:E1')
        ws.merge_cells('F1:I1')
        ws.merge_cells('J1:J2')

        # Style settings
        header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")  # Yellow (matches screenshot)
        subheader_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")  # Yellow (matches screenshot)
        section_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Light blue (matches screenshot)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Apply styles to headers
        for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=10):
            for cell in row:
                cell.fill = header_fill if cell.row == 1 else subheader_fill
                cell.font = Font(bold=True)
                cell.alignment = center_alignment
                cell.border = border

        # Define site structure (as per the screenshot)
        sites_structure = [
            ("Tower F", ["F1", "F2", "Common"]),
            ("Tower G", ["G1", "G2", "G3", "Common"]),
            ("Tower H", ["H1", "H2", "H3", "H4", "Common"])
        ]

        # Populate the table with data
        row_start = 3
        for site_group, modules in sites_structure:
            # Write the main site (e.g., Tower F) and merge cells
            ws.cell(row=row_start, column=1, value=site_group.upper())
            ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=10)
            for col in range(1, 11):
                ws.cell(row=row_start, column=col).fill = section_fill
                ws.cell(row=row_start, column=col).alignment = center_alignment
                ws.cell(row=row_start, column=col).border = border
            row_start += 1

            # Write the modules (e.g., F1, F2, Common) and populate NCR counts
            for module in modules:
                ws.cell(row=row_start, column=1, value=module)
                # Map the module to the site key used in the dictionaries
                site_key = module if module in open_ncr_data else site_group.upper()
                # Populate NCR resolved counts
                ws.cell(row=row_start, column=2, value=resolved_ncr_data[site_key]['Civil'])
                ws.cell(row=row_start, column=3, value=resolved_ncr_data[site_key]['Finishing'])
                ws.cell(row=row_start, column=4, value=resolved_ncr_data[site_key]['Works'])
                ws.cell(row=row_start, column=5, value=resolved_ncr_data[site_key]['MEP'])
                # Populate NCR open counts
                ws.cell(row=row_start, column=6, value=open_ncr_data[site_key]['Civil'])
                ws.cell(row=row_start, column=7, value=open_ncr_data[site_key]['Finishing'])
                ws.cell(row=row_start, column=8, value=open_ncr_data[site_key]['Works'])
                ws.cell(row=row_start, column=9, value=open_ncr_data[site_key]['MEP'])
                # Calculate TOTAL (sum of open NCRs)
                total = (open_ncr_data[site_key]['Civil'] + open_ncr_data[site_key]['Finishing'] +
                         open_ncr_data[site_key]['Works'] + open_ncr_data[site_key]['MEP'])
                ws.cell(row=row_start, column=10, value=total)
                # Apply styling to the row
                for col in range(1, 11):
                    ws.cell(row=row_start, column=col).fill = section_fill
                    ws.cell(row=row_start, column=col).alignment = center_alignment
                    ws.cell(row=row_start, column=col).border = border
                row_start += 1

        # Adjust column widths to match the screenshot
        ws.column_dimensions['A'].width = 15  # SITE column
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 10  # NCR columns
        ws.column_dimensions['J'].width = 8  # TOTAL column

        # Save the workbook to a BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Excel file generated successfully!")

        st.download_button(
            label="📥 Download Excel File",
            data=output,
            file_name="ncr_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Error processing JSON: {e}")

data = [
    {"Status": "Open", "Module_number": "M1", "Tower": "F", "Discipline": "FW", "Description": "Tower F, Module – 1, Fourth Floor, Flat no. 401, We have found that the main door shutter has been fixed but the shutter is damaged due to negligence in supervision.", "Created Date": "2025-04-17 00:00:00", "Expected close date": "2025-04-17 00:00:00"},
    {"Status": "Closed", "Module_number": "M2", "Tower": "G", "Discipline": "SW", "Description": "Tower - G, Module - 2, Seventh Floor, Flat No - 702, We have found that the flat main entrance electrical shaft shear wall both side concrete is bulged and offset is visible due to negligence in supervision.", "Created Date": "2025-01-24 00:00:00", "Expected close date": "2025-02-19 00:00:00"},
    {"Status": "Open", "Module_number": "M6", "Tower": "H", "Discipline": "SW", "Description": "Tower - H, Module - 6, Terrace floor, Grid - TH.53/NTX.17, We have found that the elevation column vertical reinforcement lap bars missing due to negligence in supervision. It should be rectified immediately.", "Created Date": "2024-12-10 00:00:00", "Expected close date": "2024-12-20 00:00:00"},
    {"Status": "Closed", "Module_number": "M5", "Tower": "H", "Discipline": "SW", "Description": "Tower - H, Module - 5, Shear wall - SW131, Grid - TH.42 / TH.C-D, First Floor, We have found that the expansion joint shear wall lap zone is not as per IS code due to negligence in supervision. It should be rectified.", "Created Date": "2024-12-31 00:00:00", "Expected close date": "2025-01-17 00:00:00"},
    {"Status": "Open", "Module_number": "M4", "Tower": "H", "Discipline": "SW", "Description": "Tower- H, Module – 4, Second Floor, Flat No-202 , Shear Wall No- SW04, We have found that the shear wall cover is not as per specification and wall is 40mm out from the grid due to negligence in supervision.", "Created Date": "2025-03-19 00:00:00", "Expected close date": "2025-03-26 00:00:00"},
    {"Status": "Closed", "Module_number": "M1", "Tower": "G", "Discipline": "HSE", "Description": "Tower G, Module - 1, We have found that on the Third floor steel fixing work is in progress but hard barricading around the corridor is not provided leading to violation of HSE norms due to negligence in supervision", "Created Date": "2025-03-24 00:00:00", "Expected close date": "2025-05-08 00:00:00"},
    {"Status": "Open", "Module_number": "Common_area", "Tower": "G", "Discipline": "HSE", "Description": "Following observations are found at CIPL steel yard:1) Steel bars are not stuck on concrete beam sections and lying on the ground and getting rusted.2) Steel bars after cutting and bending are not stuck on a raised platform and found lying on the ground surface.3) Spacing between two consecutive concrete beams on which steel bars are stuck on unloading are not proper due to which steel bars sag and found to be in contact of ground.Pls rectify all the above mentioned points.", "Created Date": "2025-03-24 00:00:00", "Expected close date": "2025-05-08 00:00:00"}
]