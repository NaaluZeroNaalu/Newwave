import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string





tower2 = []
tower3 = []
tower4 = []
tower5 = []
tower6 = []
tower7 = []



WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-3-70b-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KS5iR_XHOYc4N_xoId6YcXFjZR2ikINRdAyc2w2o18Oo"


def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    
    if response.status_code != 200:
        st.write(f"Failed to get access token: {response.text}")
        return None
    else:
        token_info = response.json()
        return token_info['access_token']


def generatePrompt(json_datas):
    body = {
        "input": f"""
         
        Read all data from this table carefully:
         
        {json_datas}.
        
        need a average value as percentage for green as single json  take poject name of each tower on that table
        
        Calculate the average value for green as a percentage and return the result in JSON format. Do not change the "Project" field value.

        For the "Structure" percentage, divide the green value by the non-green value.

        Use this formula:
        Structure = (Total Green / Total Non-Green) × 100

        Sample json:
       [[{{
        ""
           "Project":"Project name"
           "Tower Name:"tower name",
           "Structure":"percentage %",
           "Finishing":"0%"
        }}]

        Return the result strictly as a JSON object—no code, no explanations, only the JSON.

        Dont put <|eom_id|> or any other

        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    
    if not headers["Authorization"]:
        return "Error: No valid access token."
    
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    
    if response.status_code != 200:
        st.write(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    # st.write(json_datas)
    return response.json()['results'][0]['generated_text'].strip()




def Tower2(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower2.append(1)
                else:
                    tower2.append(0)
            else:
                tower2.append(0)

def Tower3(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['T', 'V', 'X', 'Z', 'AB', 'AD', 'AF', 'AH']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower3.append(1)
                else:
                    tower3.append(0)
            else:
                tower3.append(0)

def Tower4(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX', 'AZ', 'BB', 'BD', 'BF', 'BH', 'BJ', 'BL', 'BN', 'BP']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower4.append(1)
                else:
                    tower4.append(0)
            else:
                tower4.append(0)

def Tower5(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['DC', 'DE', 'DG', 'DI', 'DK', 'DM', 'DO', 'DQ', 'DS', 'DU', 'DW', 'DY', 'EA', 'EC']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower5.append(1)
                else:
                    tower5.append(0)
            else:
                tower5.append(0)

def Tower6(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['FK', 'FM', 'FO', 'FQ', 'FS', 'FU', 'FW', 'FY', 'GA', 'GC', 'GE', 'GG', 'GI', 'GK']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower6.append(1)
                else:
                    tower6.append(0)
            else:
                tower6.append(0)

def Tower7(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['EG', 'EI', 'EK', 'EM', 'EO', 'EQ', 'ES', 'EU', 'EW', 'EY', 'FA', 'FC', 'FE', 'FG']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower7.append(1)
                else:
                    tower7.append(0)
            else:
                tower7.append(0)


def ProcessVeridia(exceldatas, ignore_year, ignore_month):
    wb = load_workbook(exceldatas)
    sheet_name = "Revised baseline with 60d NGT"
    sheet = wb[sheet_name]

    tower2.clear()
    Tower2(sheet, ignore_year, ignore_month)
    tower3.clear()
    Tower3(sheet, ignore_year, ignore_month)
    tower4.clear()
    Tower4(sheet, ignore_year, ignore_month)
    tower5.clear()
    Tower5(sheet, ignore_year, ignore_month)
    tower6.clear()
    Tower6(sheet, ignore_year, ignore_month)
    tower7.clear()
    Tower7(sheet, ignore_year, ignore_month)

    data = {
        # "Project Name": ["VERIDIA"] * 6,
        "Tower": ["SLAB TOWER 2", "SLAB TOWER 3", "SLAB TOWER 4", "SLAB TOWER 5", "SLAB TOWER 6", "SLAB TOWER 7"],
        "Green (1)": [tower2.count(1), tower3.count(1), tower4.count(1), tower5.count(1), tower6.count(1), tower7.count(1)],
        "Non-Green (0)": [tower2.count(0), tower3.count(0), tower4.count(0), tower5.count(0), tower6.count(0), tower7.count(0)],
        
    }

    project_and_green = [{"Tower": project, "Green (1)": green} for project, green in zip(data["Tower"], data["Green (1)"])]
    json_data = json.dumps(project_and_green, indent=4)

    st.write(json_data)
    return json_data


    