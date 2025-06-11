import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from io import BytesIO

def extract_module(description, tower):
    module_pattern = r'Module\s*[-–]\s*(\d+)(?:\s*&\s*(\d+))?'
    match = re.search(module_pattern, description, re.IGNORECASE)
    if match:
        modules = []
        # Prefix with tower letter if tower is F, G, or H
        prefix = tower if tower in ['F', 'G', 'H'] else ''
        modules.append(f"{prefix}{match.group(1)}")
        if match.group(2):  # If second module is present (e.g., "6 &7")
            modules.append(f"{prefix}{match.group(2)}")
        return modules
    return None


def extract_tower(description):
    tower_pattern = r'(?:Tower\s*-?\s*[FGH]|T\s*[fgh])'
    match = re.search(tower_pattern, description, re.IGNORECASE)
    if match:
        tower = match.group(0).upper()
        if 'F' in tower:
            return 'F'
        elif 'G' in tower:
            return 'G'
        elif 'H' in tower:
            return 'H'
    return 'Common-area'


 
# Categorize disciplines
def categorize_discipline(discipline):
    if discipline in ["SW", "HSE"]:
        return "Civil"
    elif discipline == "FW":
        return "Structure Works"
    elif discipline == "EL":
        return "MEP"
    return None

def Excel_data(json_data):
    current_date = datetime(2025, 5, 29)
    threshold_date = current_date - timedelta(days=21)

    ncr_counts = {
        "F": {"Civil": 0, "Structure Works": 0, "MEP": 0, "modules": {}},
        "G": {"Civil": 0, "Structure Works": 0, "MEP": 0, "modules": {}},
        "H": {"Civil": 0, "Structure Works": 0, "MEP": 0, "modules": {}}
    }

    # Initialize modules
    for tower in ncr_counts:
        modules = set()
        for entry in json_data:
            if entry["Tower"] == tower:
                module = entry["Module count of each count"]
                modules.add(module)
        modules.add("Common")
        for module in modules:
            ncr_counts[tower]["modules"][module] = {"Civil": 0, "Structure Works": 0, "MEP": 0}

    for entry in json_data:
        expected_close = datetime.strptime(entry["Expected close date"], "%Y-%m-%d %H:%M:%S")
        if expected_close < threshold_date:
            tower = entry["Tower"]
            module = entry["Module count of each count"]
            category = categorize_discipline(entry["Discipline"])
            if category:
                ncr_counts[tower][category] += 1
                if module in ncr_counts[tower]["modules"]:
                    ncr_counts[tower]["modules"][module][category] += 1
                else:
                    for m in [m.strip() for m in module.split(",")]:
                        if m in ncr_counts[tower]["modules"]:
                            ncr_counts[tower]["modules"][m][category] += 1

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NCR Summary"

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    bold_font = Font(bold=True)

    today_date = datetime.now().strftime("%d-%m-%Y")
    ws.cell(row=1, column=1, value=f"NCR Summary Report - {today_date}")
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1).fill = yellow_fill
    ws.cell(row=1, column=1).font = bold_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row=2, column=1, value="Site")
    ws.cell(row=2, column=2, value="NCR resolved beyond 21 days")
    ws.cell(row=2, column=5, value="NCR open beyond 21 days")
    ws.cell(row=2, column=8, value="TOTAL")
    ws.cell(row=3, column=2, value="Civil")
    ws.cell(row=3, column=3, value="Structure Works")
    ws.cell(row=3, column=4, value="MEP")
    ws.cell(row=3, column=5, value="Civil")
    ws.cell(row=3, column=6, value="Structure Works")
    ws.cell(row=3, column=7, value="MEP")

    ws.merge_cells('A2:A3')
    ws.merge_cells('B2:D2')
    ws.merge_cells('E2:G2')
    ws.merge_cells('H2:H3')

    for row in [2, 3]:
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    table_data = [
        ("Tower F", "F", True),
        ("F1", "F", False),
        ("F2", "F", False),
        ("Common", "F", False),
        ("Tower G", "G", True),
        ("G1", "G", False),
        ("G2", "G", False),
        ("G3", "G", False),
        ("Common", "G", False),
        ("Tower H", "H", True),
        ("H1", "H", False),
        ("H2", "H", False),
        ("H3", "H", False),
        ("H4", "H", False),
        ("H5", "H", False),
        ("H6", "H", False),
        ("H7", "H", False),
        ("Common", "H", False)
    ]

    for row_idx, (site, tower, is_tower_header) in enumerate(table_data, 4):
        ws.cell(row=row_idx, column=1).value = site
        for col in range(2, 5):
            ws.cell(row=row_idx, column=col).value = 0

        if is_tower_header:
            civil = ncr_counts[tower]["Civil"]
            sw = ncr_counts[tower]["Structure Works"]
            mep = ncr_counts[tower]["MEP"]
        else:
            module_counts = ncr_counts[tower]["modules"].get(site, {"Civil": 0, "Structure Works": 0, "MEP": 0})
            civil = module_counts["Civil"]
            sw = module_counts["Structure Works"]
            mep = module_counts["MEP"]

        ws.cell(row=row_idx, column=5).value = civil
        ws.cell(row=row_idx, column=6).value = sw
        ws.cell(row=row_idx, column=7).value = mep
        ws.cell(row=row_idx, column=8).value = civil + sw + mep

        if is_tower_header:
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = light_blue_fill

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in range(1, len(table_data) + 4):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border

    column_widths = [15, 12, 15, 12, 12, 15, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output



json_data = [
    {"Status":"Open","Module count of each count":"G1","Tower":"G","Discipline":"HSE","Description":"Tower G, Module - 1, We have found that on the Third floor steel fixing work is in progress but hard barricading around the corridor is not provided leading to violation of HSE norms due to negligence in supervision","Created Date":"2025-03-24 00:00:00","Expected close date":"2025-05-08 00:00:00"},
    {"Status":"Open","Module count of each count":"G2","Tower":"G","Discipline":"EL","Description":"Tower - G, Module - 2, Seventh Floor, Flat No - 701, We have found that the master bedroom wall electrical conduit is exposed from wall surface due to negligence in supervision","Created Date":"2025-01-24 00:00:00","Expected close date":"2025-02-19 00:00:00"},
    {"Status":"Open","Module count of each count":"H3","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 3, First Floor, Flat No – 101, Shear Wall No - SW63, We have found that the shear wall vertical steel lap not as per IS Code due to negligence in supervision. It should be rectified immediately.","Created Date":"2025-02-18 00:00:00","Expected close date":"2025-03-05 00:00:00"},
    {"Status":"Open","Module count of each count":"G1","Tower":"G","Discipline":"SW","Description":"Tower- G, Module- 1, Second Floor, Flat No- 301, We have found that the guest bedroom shear wall cover is not as per specification (50mm instead off 25mm) due to negligence in supervision.","Created Date":"2025-03-24 00:00:00","Expected close date":"2025-05-08 00:00:00"},
    {"Status":"Open","Module count of each count":"G2","Tower":"G","Discipline":"SW","Description":"Tower - G, Module - 2, Eighth Floor, Flat No – 802, Column - C12 (Grid – TG.18 / TG.A) & C13 (Grid- TG.20/TG.A), We have found that the column cover is not as per specification (0mm instead of 40mm) and column is 30mm out from the grid line as well as steel has been majorly joggled without any prior approval due to negligence in supervision. It should be rectified as per recommendation from the structural consultant.","Created Date":"2024-12-31 00:00:00","Expected close date":"2025-02-04 00:00:00"},
    {"Status":"Open","Module count of each count":"H6, H7","Tower":"H","Discipline":"SW","Description":"Tower - H, Module – 6 & 7, Sixth Floor, We have found that the corridor area slab beam bottom not finished properly and cracks are visible at the beam bottom surface due to negligence in supervision. It should be rectified.","Created Date":"2024-12-24 00:00:00","Expected close date":"2024-12-26 00:00:00"},
    {"Status":"Open","Module count of each count":"F1","Tower":"F","Discipline":"FW","Description":"Tower F, Module – 1, Fifth Floor, Flat No - 502, We have found that the master bedroom wall putty work has been completed but crack is visible on the wall surface due to negligence in supervision","Created Date":"2025-04-17 00:00:00","Expected close date":"2025-05-05 00:00:00"},
    {"Status":"Open","Module count of each count":"H7","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 7, Eighth Floor, Flat No - 803, We have found that the guest bedroom shear wall concrete is bulged due to negligence in supervision. It should be rectified.","Created Date":"2024-12-24 00:00:00","Expected close date":"2024-12-26 00:00:00"},
    {"Status":"Open","Module count of each count":"F2","Tower":"F","Discipline":"FW","Description":"Tower F, Module 2, First to Seventh Floor, We have found that the lift area shear wall whitewashing work has been completed but yellow grease marks are visible due to bad workmanship and bad supervision. It should be rectified immediately.","Created Date":"2024-07-04 00:00:00","Expected close date":"2024-12-26 00:00:00"},
    {"Status":"Open","Module count of each count":"H6","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 6, Seventh Floor, Flat No - 701, We have found that the dining area shear wall concrete is bulged and wall is not in plumb due to negligence in supervision. It should be rectified immediately.","Created Date":"2024-12-13 00:00:00","Expected close date":"2024-12-20 00:00:00"},
    {"Status":"Open","Module count of each count":"Common","Tower":"F","Discipline":"FW","Description":"Tower F, Module 2, Third Floor, Flat no – 304, We have found that the whole flat wall putty work has been completed but yellow grease spots are visible due to negligence in supervision. It should be rectified immediately.","Created Date":"2024-09-01 00:00:00","Expected close date":"2024-09-13 00:00:00"},
    {"Status":"Open","Module count of each count":"Common","Tower":"H","Discipline":"SW","Description":"It is observed that at Tower F lift wall (Shear wall 38) offset is found.It should be rectified immediately.","Created Date":"2023-01-25 00:00:00","Expected close date":"2023-01-25 00:00:00"},
    {"Status":"Open","Module count of each count":"H2","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 2, Stilt Floor, Beam No - THB1, We have found that the slab beam is honeycombed as well as steel is exposed due to negligence in supervision. It should be rectified immediately.","Created Date":"2024-12-13 00:00:00","Expected close date":"2024-12-26 00:00:00"},
    {"Status":"Open","Module count of each count":"H6","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 6, Terrace Floor, We have found that the slab beam steel binding work has been done but cover blocks are not fixed in the beam due to negligence in supervision. It should be rectified immediately.","Created Date":"2024-12-13 00:00:00","Expected close date":"2024-12-20 00:00:00"},
    {"Status":"Open","Module count of each count":"F1","Tower":"F","Discipline":"FW","Description":"Tower F, Module – 1, Fourth Floor, Flat no. 401, We have found that the main door shutter has been fixed but the shutter is damaged due to negligence in supervision.","Created Date":"2025-04-17 00:00:00","Expected close date":"2025-04-17 00:00:00"},
    {"Status":"Open","Module count of each count":"H3","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 3, First Floor, Shear wall No – SW66, We have found that the shear wall vertical steel lap length is 405mm instead of 505mm as well as shear wall lap zone is not as per IS code due to negligence in supervision. It should be rectified.","Created Date":"2025-01-23 00:00:00","Expected close date":"2025-01-27 00:00:00"},
    {"Status":"Open","Module count of each count":"Common","Tower":"H","Discipline":"SW","Description":"Non Tower, We have found that the non tower area column footing pit PCC work is in progress but PCC has been cracked due to negligence in supervision.","Created Date":"2025-04-10 00:00:00","Expected close date":"2025-04-10 00:00:00"},
    {"Status":"Open","Module count of each count":"G2","Tower":"G","Discipline":"SW","Description":"Tower - G, Module - 2, Seventh Floor, Flat No - 702, We have found that the flat main entrance electrical shaft shear wall both side concrete is bulged and offset is visible due to negligence in supervision.","Created Date":"2025-01-24 00:00:00","Expected close date":"2025-02-19 00:00:00"},
    {"Status":"Open","Module count of each count":"G2","Tower":"G","Discipline":"SW","Description":"Tower - G, Module - 2, Eighth Floor, Flat No - 801, Column - C22 (Grid – TG.14 / TG.C) & C10 (Grid - TG.15 / TG.A), We have found that the column cover is not as per specification (0mm instead of 40mm) and column is 30mm out from the grid line as well as steel has been majorly joggled without any prior approval due to negligence in supervision. It should be rectified as per recommendation from the structural consultant.","Created Date":"2024-12-31 00:00:00","Expected close date":"2025-02-04 00:00:00"},
    {"Status":"Open","Module count of each count":"H6","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 6, Terrace floor, Grid - TH.53/NTX.17, We have found that the elevation column vertical reinforcement lap bars missing due to negligence in supervision. It should be rectified immediately.","Created Date":"2024-12-10 00:00:00","Expected close date":"2024-12-20 00:00:00"},
    {"Status":"Open","Module count of each count":"H4","Tower":"H","Discipline":"HSE","Description":"Tower H, Module 4, First Floor Level, We have found that the housekeeping at first floor level is very poor as well as steel scrap is not stacked properly due to negligence in supervision.","Created Date":"2024-10-07 00:00:00","Expected close date":"2024-12-06 00:00:00"},
    {"Status":"Open","Module count of each count":"G2","Tower":"G","Discipline":"SW","Description":"Tower - G, Module - 2, Seventh Floor, Flat No - 701, We have found that the master bedroom toilet wall plumbing groove is not in alignment due to negligence in supervision.","Created Date":"2025-01-24 00:00:00","Expected close date":"2025-02-19 00:00:00"},
    {"Status":"Open","Module count of each count":"F2","Tower":"F","Discipline":"FW","Description":"Tower F, Module 2, Fourth Floor, Flat No- 401 & 402, We have found that the wall putty work has been completed but the whole flat grease spots are visible on the wall surface due to bad workmanship and bad supervision. It should be rectified.","Created Date":"2024-10-09 00:00:00","Expected close date":"2024-10-25 00:00:00"},
    {"Status":"Open","Module count of each count":"G2","Tower":"G","Discipline":"SW","Description":"Tower - G, Module - 2, Fourth Floor to Sixth Floor, We have found that the staircase mid-landing non-structural wall has been casted but cracks are visible on the wall surface due to negligence in supervision. It should be rectified.","Created Date":"2024-12-31 00:00:00","Expected close date":"2025-03-05 00:00:00"},
    {"Status":"Open","Module count of each count":"H3","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 3, First Floor, Flat No – 101, Grid – TH.24-Th.26/TH.E, Shear Wall No - SW54, We have found that the kitchen shear wall vertical steel lap not as per IS Code due to negligence in supervision. It should be rectified.","Created Date":"2025-02-18 00:00:00","Expected close date":"2025-03-05 00:00:00"},
    {"Status":"Open","Module count of each count":"H3","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 3, First Floor, Flat No – 101, Shear Wall No - SW20, We have found that the Guest bedroom outer shear wall vertical lap not as per IS Code due to negligence in supervision.","Created Date":"2025-02-18 00:00:00","Expected close date":"2025-03-05 00:00:00"},
    {"Status":"Open","Module count of each count":"G2","Tower":"G","Discipline":"SW","Description":"Tower - G, Module - 2, Seventh Floor, Corridor Column, We have found that the column concrete is bulged and offset is visible due to negligence in supervision.","Created Date":"2025-01-24 00:00:00","Expected close date":"2025-02-19 00:00:00"},
    {"Status":"Open","Module count of each count":"H6","Tower":"H","Discipline":"HSE","Description":"Tower - H, Module - 6, Seventh Floor, Flat No - 701, We have found that the flat balcony outer barricading is not provided leading to violence of HSE norms due to negligence in supervision.","Created Date":"2024-12-13 00:00:00","Expected close date":"2024-12-20 00:00:00"},
    {"Status":"Open","Module count of each count":"H6","Tower":"H","Discipline":"SW","Description":"Tower H, Module - 6, Seventh Floor, Flat No - 701 & 702, Shear wall - SW145, SW144, SW143 & SW139, We have found that the Shear walls reinforcement has been majorly joggled without prior approval due to negligence in supervision. It should be rectified as per recommendations from structural consultant.","Created Date":"2024-10-18 00:00:00","Expected close date":"2024-10-24 00:00:00"},
    {"Status":"Open","Module count of each count":"Common","Tower":"H","Discipline":"SW","Description":"It is observed that at Tower (F) foundation level Column C10 and shear wall SW21 is found to be out of plumb.It should be rectified immediately.","Created Date":"2022-12-27 00:00:00","Expected close date":"2022-12-27 00:00:00"},
    {"Status":"Open","Module count of each count":"H4","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - F4, Column - C35 (TH.41/TH.F) & C40 (TH.41/TH.H), First Floor, We have found that the expansion joint columns casted at stilt floor level but their vertical reinforcement is missing at the first-floor level as well as all vertical bars are rebar (C35 & C40 – 16mm,20 nos. & 12mm, 8 nos.) at first floor due to negligence in supervision. Please attach Structural Consultant approval for NCR compliance.","Created Date":"2024-12-26 00:00:00","Expected close date":"2025-01-17 00:00:00"},
    {"Status":"Open","Module count of each count":"G1","Tower":"G","Discipline":"SW","Description":"Tower G, Module - 1, Second to Third Floor, We have found that the staircase mid landing beam cover is not as per drawing ( 50mm instead of 25mm ) due to negligence in supervision.","Created Date":"2025-04-10 00:00:00","Expected close date":"2025-05-08 00:00:00"},
    {"Status":"Open","Module count of each count":"H3","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 3, Stilt Floor, Grid – TH.21/TH.M to TH.K, We have found that the shear wall concrete is bulged and offset is visible due to negligence in supervision.","Created Date":"2025-01-23 00:00:00","Expected close date":"2025-01-27 00:00:00"},
    {"Status":"Open","Module count of each count":"F2","Tower":"F","Discipline":"FW","Description":"Tower F, Module 2, Fourth Floor, Flat no – 404, We have found that the whole flat wall putty work has been completed but yellow grease spots are visible due to negligence in supervision. It should be rectified immediately.","Created Date":"2024-09-01 00:00:00","Expected close date":"2024-09-13 00:00:00"},
    {"Status":"Open","Module count of each count":"H7","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 7, Eight Floor, Flat No - 801, We have found that the guest bedroom non-structural wall concrete is bulged and electrical module box is out from the wall surface due to negligence in supervision. It should be rectified.","Created Date":"2024-12-24 00:00:00","Expected close date":"2024-12-26 00:00:00"},
    {"Status":"Open","Module count of each count":"Common","Tower":"H","Discipline":"SW","Description":"It is observed that at CIPL QC lab temperature of cubes curing tank is not maintained which affects the strength of Cubes.Al Penal Code of India (Amendment) Act, 2013, section 376D makes it mandatory to take victim’s consent in writing before subjecting her to a medical examination in rape cases. Also the thermometer availed at lab is not appropriate.Pls arrange for a digital thermometer and a thermostat to maintain water temperature.","Created Date":"2022-12-29 00:00:00","Expected close date":"2024-03-11 00:00:00"},
    {"Status":"Open","Module count of each count":"H2","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 2, Stilt Floor, Column- C57, C58, C59 & C60, We have found that the column 30mm thick starter was made with cement mortar instead of cement concrete due to negligence in supervision. It should be rectified immediately.","Created Date":"2024-12-13 00:00:00","Expected close date":"2024-12-26 00:00:00"},
    {"Status":"Open","Module count of each count":"H4","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 4, Stilt Floor, We have found that the slab undulation is visible and slab is not in level at the bottom surface near the staircase due to negligence in supervision.","Created Date":"2025-01-23 00:00:00","Expected close date":"2025-02-17 00:00:00"},
    {"Status":"Open","Module count of each count":"H2","Tower":"H","Discipline":"HSE","Description":"Tower - H, Module - 2, Stilt Floor, We have found that the labours are working on scaffolding without wearing safety belt leading to violence of HSE norms due to negligence in supervision. It should be rectified immediately.","Created Date":"2024-12-13 00:00:00","Expected close date":"2024-12-26 00:00:00"},
    {"Status":"Open","Module count of each count":"G2","Tower":"G","Discipline":"SW","Description":"Tower - G, Module - 2, Seventh Floor, Flat No - 701, We have found that the kitchen back side shear wall repairing work has been completed but wall is still not in plumb and offset of 25mm is visible due to negligence in supervision.","Created Date":"2025-01-24 00:00:00","Expected close date":"2025-02-19 00:00:00"},
    {"Status":"Open","Module count of each count":"G1","Tower":"G","Discipline":"SW","Description":"Tower- G, Module-1, Third Floor, Shear Wall No- SW38, Flat No-301, We have found that the shear wall cover not as per specification (50mm instead of 25mm) due to negligence in supervision.","Created Date":"2025-03-24 00:00:00","Expected close date":"2025-05-08 00:00:00"},
    {"Status":"Open","Module count of each count":"H5","Tower":"H","Discipline":"SW","Description":"Tower H, Module - 5, Fourth floor, Flat No- 402 , Shear Wall No - SW157, We have found that the shear wall cover (10mm instead of 25mm) is not as per specification as well as vertical reinforcement has been joggled without any approval. It should be rectified.","Created Date":"2025-04-07 00:00:00","Expected close date":"2025-04-21 00:00:00"},
    {"Status":"Open","Module count of each count":"H4","Tower":"H","Discipline":"SW","Description":"Tower- H, Module – 4, Second Floor, Flat No-202 , Shear Wall No- SW04, We have found that the shear wall cover is not as per specification and wall is 40mm out from the grid due to negligence in supervision.","Created Date":"2025-03-19 00:00:00","Expected close date":"2025-03-26 00:00:00"},
    {"Status":"Open","Module count of each count":"H7","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 7, Eighth Floor, Flat No – 802, We have found that the other bedroom shear wall has been casted but crack is visible on the wall surface due to negligence in supervision. It should be rectified.","Created Date":"2025-01-02 00:00:00","Expected close date":"2025-01-06 00:00:00"},
    {"Status":"Open","Module count of each count":"H6","Tower":"H","Discipline":"SW","Description":"Tower H, Module - 6, Sixth Floor, Flat No-602, We have found that the dining area wall electrical module box is not in alignment due to negligence in supervision.","Created Date":"2024-10-22 00:00:00","Expected close date":"2024-10-24 00:00:00"},
    {"Status":"Open","Module count of each count":"Common","Tower":"H","Discipline":"SW","Description":"Following observations are found at CIPL steel yard:1) Steel bars are not stuck on concrete beam sections and lying on the ground and getting rusted.2) Steel bars after cutting and bending are not stuck on a raised platform and found lying on the ground surface.3) Spacing between two consecutive concrete beams on which steel bars are stuck on unloading are not proper due to which steel bars sag and found to be in contact of ground.Pls rectify all the above mentioned points.","Created Date":"2022-12-29 00:00:00","Expected close date":"2023-04-05 00:00:00"},
    {"Status":"Open","Module count of each count":"H6","Tower":"H","Discipline":"SW","Description":"Tower H, Module - 6, Seventh Floor, Flat No - 701,702 & 703, Column No - C83, C82, C79 & C119, We have found that the columns reinforcement has been majorly joggled without prior approval due to negligence in supervision. It should be rectified as per recommendations from structural consultant.","Created Date":"2024-10-18 00:00:00","Expected close date":"2024-10-24 00:00:00"},
    {"Status":"Open","Module count of each count":"H6, H7","Tower":"H","Discipline":"SW","Description":"Tower - H, Module – 6 & 7, Fourth floor to Sixth Floor, We have found that the toilet outer wall construction joint undulation and offset are visible due to negligence in supervision. It should be rectified.","Created Date":"2024-12-24 00:00:00","Expected close date":"2024-12-26 00:00:00"},
    {"Status":"Open","Module count of each count":"H7","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 7, Eighth Floor, Flat No – 804, We have found that the guest bedroom plumbing shaft nonstructural wall concrete is bulged as well as wooden piece not removed due to negligence in supervision.","Created Date":"2025-01-02 00:00:00","Expected close date":"2025-01-06 00:00:00"},
    {"Status":"Open","Module count of each count":"F1","Tower":"F","Discipline":"FW","Description":"Tower F, Module – 1, Fifth Floor, Flat No - 501, We have found that the kitchen wall tile fixing work has been completed but wall tile fixed is not in plumb due to negligence in supervision.","Created Date":"2025-04-17 00:00:00","Expected close date":"2025-05-20 00:00:00"},
    {"Status":"Open","Module count of each count":"G2","Tower":"G","Discipline":"SW","Description":"Tower G, Module - 2,  2nd to 3rd Floor, We have found that the guest bedroom toilet outer non-structural wall joint has been cracked as well as wall joint is not finished properly due to negligence in supervision.","Created Date":"2025-04-10 00:00:00","Expected close date":"2025-05-18 00:00:00"},
    {"Status":"Open","Module count of each count":"G1","Tower":"G","Discipline":"SW","Description":"Tower- G, Module-1, First Floor, Flat No – 102, We have found that the kitchen nonstructural wall concrete is bulged as well as wall not in plumb due to negligence in supervision.","Created Date":"2025-03-11 00:00:00","Expected close date":"2025-04-02 00:00:00"},
    {"Status":"Open","Module count of each count":"Common","Tower":"G","Discipline":"SW","Description":"NTA Beam side at Grid TF.21/TF.A to TF.B, Bulging and honeycomb has been observed . it should be rectified immediately","Created Date":"2023-02-21 00:00:00","Expected close date":"2023-02-21 00:00:00"},
    {"Status":"Open","Module count of each count":"H5","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 5, Shear wall - SW131, Grid - TH.42 / TH.C-D, First Floor, We have found that the expansion joint shear wall lap zone is not as per IS code due to negligence in supervision. It should be rectified.","Created Date":"2024-12-31 00:00:00","Expected close date":"2025-01-17 00:00:00"},
    {"Status":"Open","Module count of each count":"H3","Tower":"H","Discipline":"SW","Description":"Tower - H, Module - 3, First Floor, Flat No – 104, We have found that the guest toilet non structural wall mivan plate has been fixed without proper layout marking due to negligence in supervision.","Created Date":"2025-03-05 00:00:00","Expected close date":"2025-05-08 00:00:00"},
    {"Status":"Open","Module count of each count":"F1","Tower":"F","Discipline":"FW","Description":"Tower F, Module – 1, Fifth Floor, Flat No - 501, We have found that the L balcony outer side wall putty work has been completed but grease spot is visible on the wall surface due to negligence in supervision","Created Date":"2025-04-17 00:00:00","Expected close date":"2025-04-17 00:00:00"}
]


# # Function to transform the JSON
# def transform_json(input_data):
#     output = []
    
#     # Process "NCR open beyond 21 days"
#     ncr_open = input_data.get("NCR open beyond 21 days", {}).get("Sites", {}).get("Common_Area", {})
#     if ncr_open:
#         descriptions = ncr_open.get("Descriptions", [])
#         created_dates = ncr_open.get("Created Date (WET)", [])
#         expected_close_dates = ncr_open.get("Expected Close Date (WET)", [])
#         statuses = ncr_open.get("Status", [])
#         disciplines = ncr_open.get("Discipline", [])
#         modules = ncr_open.get("Modules", [])
        
#         for i in range(len(descriptions)):
#             tower = extract_tower(descriptions[i])
#             # Try to extract module from description first
#             extracted_modules = extract_module(descriptions[i], tower)
#             # Use extracted modules if available, otherwise fall back to provided modules
#             module_list = extracted_modules if extracted_modules else modules[i] if i < len(modules) else ["Unknown"]
#             # Adjust module names for fallback modules if tower is F, G, or H
#             if not extracted_modules and tower in ['F', 'G', 'H']:
#                 module_list = [f"{tower}{m[1:]}" if m.startswith('M') else m for m in module_list]
#             module = ", ".join(module_list)
#             record = {
#                 "Status": statuses[i] if i < len(statuses) else "Unknown",
#                 "Module count of each count": module,
#                 "Tower": tower,
#                 "Discipline": disciplines[i] if i < len(disciplines) else "Unknown",
#                 "Description": descriptions[i].strip(),
#                 "Created Date": created_dates[i] if i < len(created_dates) else "Unknown",
#                 "Expected close date": expected_close_dates[i] if i < len(expected_close_dates) else "Unknown"
#             }
#             output.append(record)
    
#     # Process Veridia-Tower sections
#     for section in ["Veridia-Tower-403-CommonArea", "Veridia-Tower-20", "Veridia-Tower-404-CommonArea"]:
#         section_data = input_data.get(section, {})
#         descriptions = section_data.get("Descriptions", [])
#         created_dates = section_data.get("Created Date (WET)", [])
#         expected_close_dates = section_data.get("Expected Close Date (WET)", [])
#         statuses = section_data.get("Status", [])
#         disciplines = section_data.get("Discipline", [])
#         modules = section_data.get("Modules", [])
        
#         for i in range(len(descriptions)):
#             tower = extract_tower(descriptions[i])
#             # Try to extract module from description first
#             extracted_modules = extract_module(descriptions[i], tower)
#             # Use extracted modules if available, otherwise fall back to provided modules
#             module_list = extracted_modules if extracted_modules else modules[i] if i < len(modules) else ["Unknown"]
#             # Adjust module names for fallback modules if tower is F, G, or H
#             if not extracted_modules and tower in ['F', 'G', 'H']:
#                 module_list = [f"{tower}{m[1:]}" if m.startswith('M') else m for m in module_list]
#             module = ", ".join(module_list)
#             record = {
#                 "Status": statuses[i] if i < len(statuses) else "Unknown",
#                 "Module count of each count": module,
#                 "Tower": tower,
#                 "Discipline": disciplines[i] if i < len(disciplines) else "Unknown",
#                 "Description": descriptions[i].strip(),
#                 "Created Date": created_dates[i] if i < len(created_dates) else "Unknown",
#                 "Expected close date": expected_close_dates[i] if i < len(expected_close_dates) else "Unknown"
#             }
#             output.append(record)
    
#     return output



if st.button("Generate Report"):
    excel_file = Excel_data(json_data)

    st.success("Excel file is ready!")

    st.download_button(
        label="📥 Download Excel Report",
        data=excel_file,
        file_name="ncr_summary.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# Transform the JSON and output to a file
# transformed_data = transform_json(input_json)
# with open('transformed_ncr.json', 'w') as f:
#     json.dump(transformed_data, f, indent=4)

# # Print a sample of the transformed JSON
# print(json.dumps(transformed_data, indent=4))  # Print first two records as a sample

