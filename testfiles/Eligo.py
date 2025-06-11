import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re
import io
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from io import BytesIO

def extract_module(description, tower):
    module_pattern = r'Module\s*[-–]\s*(\d+)(?:\s*&\s*(\d+))?'
    match = re.search(module_pattern, description, re.IGNORECASE)
    if match:
        modules = []
        # Prefix with tower letter if tower is F, G, or H
        prefix = tower if tower in ['F', 'G', 'H'] else ''
        modules.append(f"{prefix}{match.group(1)}")
        if match.group(2):  # If second module is present (e.g., "6 &7")
            modules.append(f"{prefix}{match.group(2)}")
        return modules
    return None


def extract_tower(description):
    tower_pattern = r'(?:Tower\s*-?\s*[FGH]|T\s*[fgh])'
    match = re.search(tower_pattern, description, re.IGNORECASE)
    if match:
        tower = match.group(0).upper()
        if 'F' in tower:
            return 'F'
        elif 'G' in tower:
            return 'G'
        elif 'H' in tower:
            return 'H'
    return 'Common-area'


 
# Categorize disciplines
def categorize_discipline(discipline):
    if discipline in ["SW", "HSE"]:
        return "Civil"
    elif discipline == "FW":
        return "Structure Works"
    elif discipline == "EL":
        return "MEP"
    return None

def Excel_data(json_data):
    current_date = datetime(2025, 5, 29)
    threshold_date = current_date - timedelta(days=21)

    ncr_counts = {
        "F": {"Civil": 0, "Structure Works": 0, "MEP": 0, "modules": {}},
        "G": {"Civil": 0, "Structure Works": 0, "MEP": 0, "modules": {}},
        "H": {"Civil": 0, "Structure Works": 0, "MEP": 0, "modules": {}}
    }

    # Initialize modules
    for tower in ncr_counts:
        modules = set()
        for entry in json_data:
            if entry["Tower"] == tower:
                module = entry["Module count of each count"]
                modules.add(module)
        modules.add("Common")
        for module in modules:
            ncr_counts[tower]["modules"][module] = {"Civil": 0, "Structure Works": 0, "MEP": 0}

    for entry in json_data:
        expected_close = datetime.strptime(entry["Expected close date"], "%Y-%m-%d %H:%M:%S")
        if expected_close < threshold_date:
            tower = entry["Tower"]
            module = entry["Module count of each count"]
            category = categorize_discipline(entry["Discipline"])
            if category:
                ncr_counts[tower][category] += 1
                if module in ncr_counts[tower]["modules"]:
                    ncr_counts[tower]["modules"][module][category] += 1
                else:
                    for m in [m.strip() for m in module.split(",")]:
                        if m in ncr_counts[tower]["modules"]:
                            ncr_counts[tower]["modules"][m][category] += 1

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NCR Summary"

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    bold_font = Font(bold=True)

    today_date = datetime.now().strftime("%d-%m-%Y")
    ws.cell(row=1, column=1, value=f"NCR Summary Report - {today_date}")
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1).fill = yellow_fill
    ws.cell(row=1, column=1).font = bold_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row=2, column=1, value="Site")
    ws.cell(row=2, column=2, value="NCR resolved beyond 21 days")
    ws.cell(row=2, column=5, value="NCR open beyond 21 days")
    ws.cell(row=2, column=8, value="TOTAL")
    ws.cell(row=3, column=2, value="Civil")
    ws.cell(row=3, column=3, value="Structure Works")
    ws.cell(row=3, column=4, value="MEP")
    ws.cell(row=3, column=5, value="Civil")
    ws.cell(row=3, column=6, value="Structure Works")
    ws.cell(row=3, column=7, value="MEP")

    ws.merge_cells('A2:A3')
    ws.merge_cells('B2:D2')
    ws.merge_cells('E2:G2')
    ws.merge_cells('H2:H3')

    for row in [2, 3]:
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    table_data = [
        ("Tower F", "F", True),
        ("F1", "F", False),
        ("F2", "F", False),
        ("Common", "F", False),
        ("Tower G", "G", True),
        ("G1", "G", False),
        ("G2", "G", False),
        ("G3", "G", False),
        ("Common", "G", False),
        ("Tower H", "H", True),
        ("H1", "H", False),
        ("H2", "H", False),
        ("H3", "H", False),
        ("H4", "H", False),
        ("H5", "H", False),
        ("H6", "H", False),
        ("H7", "H", False),
        ("Common", "H", False)
    ]

    for row_idx, (site, tower, is_tower_header) in enumerate(table_data, 4):
        ws.cell(row=row_idx, column=1).value = site
        for col in range(2, 5):
            ws.cell(row=row_idx, column=col).value = 0

        if is_tower_header:
            civil = ncr_counts[tower]["Civil"]
            sw = ncr_counts[tower]["Structure Works"]
            mep = ncr_counts[tower]["MEP"]
        else:
            module_counts = ncr_counts[tower]["modules"].get(site, {"Civil": 0, "Structure Works": 0, "MEP": 0})
            civil = module_counts["Civil"]
            sw = module_counts["Structure Works"]
            mep = module_counts["MEP"]

        ws.cell(row=row_idx, column=5).value = civil
        ws.cell(row=row_idx, column=6).value = sw
        ws.cell(row=row_idx, column=7).value = mep
        ws.cell(row=row_idx, column=8).value = civil + sw + mep

        if is_tower_header:
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = light_blue_fill

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in range(1, len(table_data) + 4):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border

    column_widths = [15, 12, 15, 12, 12, 15, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def combine_excel_files(excel_open, excel_closed, excel_output):
    combined_excel = BytesIO()
    
    with pd.ExcelWriter(combined_excel, engine='openpyxl') as writer:
        # Process each BytesIO object
        if excel_open:
            excel_open.seek(0)
            df_open = pd.read_excel(excel_open)
            df_open.to_excel(writer, sheet_name='Housekeeping_Open', index=False)
        
        if excel_closed:
            excel_closed.seek(0)
            df_closed = pd.read_excel(excel_closed)
            df_closed.to_excel(writer, sheet_name='Safety_Closed', index=False)
        
        if excel_output:
            excel_output.seek(0)
            df_summary = pd.read_excel(excel_output)
            df_summary.to_excel(writer, sheet_name='NCR_Summary', index=False)
    
    combined_excel.seek(0)
    return combined_excel.getvalue()


# Function to transform the JSON
def transform_json(input_data):
    output = []
    
    # Process "NCR open beyond 21 days"
    ncr_open = input_data.get("NCR open beyond 21 days", {}).get("Sites", {}).get("Common_Area", {})
    if ncr_open:
        descriptions = ncr_open.get("Descriptions", [])
        created_dates = ncr_open.get("Created Date (WET)", [])
        expected_close_dates = ncr_open.get("Expected Close Date (WET)", [])
        statuses = ncr_open.get("Status", [])
        disciplines = ncr_open.get("Discipline", [])
        modules = ncr_open.get("Modules", [])
        
        for i in range(len(descriptions)):
            tower = extract_tower(descriptions[i])
            # Try to extract module from description first
            extracted_modules = extract_module(descriptions[i], tower)
            # Use extracted modules if available, otherwise fall back to provided modules
            module_list = extracted_modules if extracted_modules else modules[i] if i < len(modules) else ["Unknown"]
            # Adjust module names for fallback modules if tower is F, G, or H
            if not extracted_modules and tower in ['F', 'G', 'H']:
                module_list = [f"{tower}{m[1:]}" if m.startswith('M') else m for m in module_list]
            module = ", ".join(module_list)
            record = {
                "Status": statuses[i] if i < len(statuses) else "Unknown",
                "Module count of each count": module,
                "Tower": tower,
                "Discipline": disciplines[i] if i < len(disciplines) else "Unknown",
                "Description": descriptions[i].strip(),
                "Created Date": created_dates[i] if i < len(created_dates) else "Unknown",
                "Expected close date": expected_close_dates[i] if i < len(expected_close_dates) else "Unknown"
            }
            output.append(record)
    
    # Process Veridia-Tower sections
    for section in ["Veridia-Tower-403-CommonArea", "Veridia-Tower-20", "Veridia-Tower-404-CommonArea"]:
        section_data = input_data.get(section, {})
        descriptions = section_data.get("Descriptions", [])
        created_dates = section_data.get("Created Date (WET)", [])
        expected_close_dates = section_data.get("Expected Close Date (WET)", [])
        statuses = section_data.get("Status", [])
        disciplines = section_data.get("Discipline", [])
        modules = section_data.get("Modules", [])
        
        for i in range(len(descriptions)):
            tower = extract_tower(descriptions[i])
            # Try to extract module from description first
            extracted_modules = extract_module(descriptions[i], tower)
            # Use extracted modules if available, otherwise fall back to provided modules
            module_list = extracted_modules if extracted_modules else modules[i] if i < len(modules) else ["Unknown"]
            # Adjust module names for fallback modules if tower is F, G, or H
            if not extracted_modules and tower in ['F', 'G', 'H']:
                module_list = [f"{tower}{m[1:]}" if m.startswith('M') else m for m in module_list]
            module = ", ".join(module_list)
            record = {
                "Status": statuses[i] if i < len(statuses) else "Unknown",
                "Module count of each count": module,
                "Tower": tower,
                "Discipline": disciplines[i] if i < len(disciplines) else "Unknown",
                "Description": descriptions[i].strip(),
                "Created Date": created_dates[i] if i < len(created_dates) else "Unknown",
                "Expected close date": expected_close_dates[i] if i < len(expected_close_dates) else "Unknown"
            }
            output.append(record)
    
    return output







# if st.button("Generate Report"):
#     excel_file = Excel_data(json_data)

#     st.success("Excel file is ready!")

#     st.download_button(
#         label="📥 Download Excel Report",
#         data=excel_file,
#         file_name="ncr_summary.xlsx",
#         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
# Transform the JSON and output to a file
# transformed_data = transform_json(input_json)
# with open('transformed_ncr.json', 'w') as f:
#     json.dump(transformed_data, f, indent=4)

# # Print a sample of the transformed JSON
# print(json.dumps(transformed_data, indent=4))  # Print first two records as a sample

