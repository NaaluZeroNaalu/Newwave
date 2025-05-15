import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from EWS_LIG import *
from Tower_G_and_H import *
from Veridia import *
from Finishing import *
from Eden import *
from datetime import date
from dateutil.relativedelta import relativedelta


if 'tower2_finishing' not in st.session_state:
    st.session_state.tower2_finishing = 0
if 'tower3_finishing' not in st.session_state:
    st.session_state.tower3_finishing = 0
if 'tower4_finishing' not in st.session_state:
    st.session_state.tower4_finishing = 0
if 'tower5_finishing' not in st.session_state:
    st.session_state.tower5_finishing = 0
if 'tower6_finishing' not in st.session_state:
    st.session_state.tower6_finishing = 0
if 'tower7_finishing' not in st.session_state:
    st.session_state.tower7_finishing = 0


if 'towerf_finishing' not in st.session_state:
    st.session_state.towerf_finishing = 0
if 'towerg_finishing' not in st.session_state:
    st.session_state.towerg_finishing = 0
if 'towerh_finishing' not in st.session_state:
    st.session_state.towerh_finishing = 0

COS_API_KEY = "ehl6KMyT95fwzKf7sPW_X3eKFppy_24xbm4P1Yk-jqyU"
COS_SERVICE_INSTANCE_ID = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:projectreportnew"
COS_ENDPOINT = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
COS_BUCKET = "projectreportnew"


def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Write the DataFrame to Excel, starting from row 1 to leave space for the title
        df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=1)
        
        # Get the xlsxwriter workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Define a format for the title row with yellow background
        title_format = workbook.add_format({
            'bold': True,
            'bg_color': 'yellow',
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # Add a title in the first row (e.g., "Tower Project Status")
        worksheet.write(0, 0, f'Overall Project Report ({date.today()})', title_format)
        
        # Merge cells across the first row for the title (assuming the DataFrame has columns)
        worksheet.merge_range(0, 0, 0, len(df.columns)-1, 'Tower Project Status', title_format)
        
    return output.getvalue()


cos_client = ibm_boto3.client(
    's3',
    ibm_api_key_id=COS_API_KEY,
    ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
    config=Config(signature_version='oauth'),
    endpoint_url=COS_ENDPOINT
)


def get_cos_files():
    try:
        response = cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            print("No .json files found in the bucket 'ozonetell'. Please ensure JSON files are uploaded.")
        return files
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        return []
    

files = get_cos_files()
st.write(files)


today = date.today()
prev_month = today - relativedelta(months=1)


month_year = today.strftime("%m-%Y")
prev_month_year = prev_month.strftime("%m-%Y")

st.write("This month:", month_year)
st.write("Previous month:", prev_month_year)


# ========EWS LIG
for file in files:
   
    try:
        if file.startswith("EWS LIG"):
            if "Structure Work Tracker" in file:
                if month_year in file:
                    st.write("✅ Current month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    ews_lig = ProcessEWSLIG(io.BytesIO(response['Body'].read()))
                elif prev_month_year in file:
                    st.write("🕓 Previous month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    ews_lig = ProcessEWSLIG(io.BytesIO(response['Body'].read()))
                    st.write(ews_lig)
    except Exception as e:
        st.error(e)
# ========EWS LIG




for file in files:
    
    try:
       if file.startswith("Eligo"):
           if "Tower G Finishing Tracker" in file:
            
                if month_year in file:
                    st.write("✅ Current month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    GetTowerGFinishing(io.BytesIO(response['Body'].read()))
                    
                    break
                elif prev_month_year in file:
                    st.write("🕓 Previous month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    GetTowerGFinishing(io.BytesIO(response['Body'].read()))
                    break
    except Exception as e:
       st.info(e)
                

for file in files:
   
    try:
        if file.startswith("Eligo"):
            if "Tower H Finishing Tracker" in file:
                if month_year in file:
                    st.write("✅ Current month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    GetTowerHFinishing(io.BytesIO(response['Body'].read()))
                    
                    break
                elif prev_month_year in file:
                    st.write("🕓 Previous month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    eligo = GetTowerHFinishing(io.BytesIO(response['Body'].read()))
                    break
    except Exception as e:
        st.info(e)
                

#========ELIGO
for file in files:
   
    try:
        if file.startswith("Eligo"):
            if "Structure Work Tracker" in file:
                if month_year in file:
                    st.write("✅ Current month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    eligo = ProcessGandH(io.BytesIO(response['Body'].read()))
                    st.write(eligo)
                elif prev_month_year in file:
                    st.write("🕓 Previous month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    eligo = ProcessGandH(io.BytesIO(response['Body'].read()))
                    st.write(eligo)
    except Exception as e:
        st.info(e)
#========ELIGO



#================EDEN

for file in files:
   
    try:
        if file.startswith("Eden"):
            if "Structure Work Tracker" in file:
                if month_year in file:
                    st.write("✅ Current month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    Eden = get_percentages(io.BytesIO(response['Body'].read()))
                    st.write(Eden)
                elif prev_month_year in file:
                    st.write("🕓 Previous month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    Eden =  get_percentages(io.BytesIO(response['Body'].read()))
                    st.write(Eden)
    except Exception as e:
        st.info(e)

#================EDEN



#============veridia
for file in files:
   
    try:
        if file.startswith("Veridia"):
            if "Tower 4 Finishing Tracker" in file:
                if month_year in file:
                    st.write("✅ Current month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    GetTower4Finishing(io.BytesIO(response['Body'].read()))
                elif prev_month_year in file:
                    st.write("🕓 Previous month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    GetTower4Finishing(io.BytesIO(response['Body'].read()))
    except Exception as e:
        st.info(e)

for file in files:
   
    try:
        if file.startswith("Veridia"):
            if "Tower 5 Finishing Tracker" in file:
                if month_year in file:
                    st.write("✅ Current month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    GetTower5Finishing(io.BytesIO(response['Body'].read()))
                elif prev_month_year in file:
                    st.write("🕓 Previous month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    GetTower5Finishing(io.BytesIO(response['Body'].read()))
    except Exception as e:
        st.info(e)

for file in files:
                
    try:
        if file.startswith("Veridia"):
            if "Structure Work Tracker" in file:
                if month_year in file:
                    st.write("✅ Current month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    veridia = ProcessVeridia(io.BytesIO(response['Body'].read()))
                    st.write(veridia)
                elif prev_month_year in file:
                    st.write("🕓 Previous month:", file)
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    veridia = ProcessVeridia(io.BytesIO(response['Body'].read()))
                    st.write(veridia)
    except Exception as e:
        st.info(e)
#============veridia

combined_data = ews_lig + veridia + eligo + Eden


df = pd.DataFrame(combined_data)
excel_data = to_excel(df)
st.session_state.overall = excel_data
st.title("Tower Project Status Table")
st.session_state.overalldf = df
# Display the dataframe as a table
st.dataframe(df)

st.download_button(
    label="Download as Excel",
    data=excel_data,
    file_name="Overall_Project_Report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)



