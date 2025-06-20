import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string
from dateutil.relativedelta import relativedelta
from datetime import date
from milestone.veridia import *
from milestone.EwsLig import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from milestone.Eligo import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from milestone.Tower4 import *
from milestone.Tower5 import *


def process_activity_data(data):
    """Process JSON data and create activity count by month"""
    df = pd.DataFrame(data)
    
    # Convert finish date to datetime
    df['Finish_Date'] = pd.to_datetime(df['Finish'])
    
    # Create a pivot table to count activities by name and month
    pivot_df = df.groupby(['Activity Name', 'Finish Month']).size().reset_index(name='Count')
    
    # Create the final dataframe in the desired format
    activity_summary = pivot_df.pivot(index='Activity Name', columns='Finish Month', values='Count').fillna(0)
    
    # Convert to integers
    activity_summary = activity_summary.astype(int)
    
    # Reset index to make Activity Name a column
    activity_summary = activity_summary.reset_index()
    
    return activity_summary

def create_tower_format(activity_summary, tower_name="Tower-ABC"):
    """Create the Tower format similar to your image"""
    # Create milestone rows
    result_data = []
    
    # Add tower header
    tower_row = [tower_name, "Activity Name"] + list(activity_summary.columns[1:])
    result_data.append(tower_row)
    
    # Add each activity as a milestone
    for idx, row in activity_summary.iterrows():
        milestone_name = f"Milestone-{idx + 1}"
        activity_name = row['Activity Name']
        counts = [f"{int(count)} activities" if count > 0 else "No activities" for count in row[1:]]
        
        milestone_row = [milestone_name, activity_name] + counts
        result_data.append(milestone_row)
    
    # Create DataFrame
    columns = ["Tower", "Activity Name"] + [f"Month {i+1}" for i in range(len(activity_summary.columns)-1)]
    final_df = pd.DataFrame(result_data[1:], columns=result_data[0])
    
    return final_df

def to_excel(df):
    """Convert DataFrame to Excel bytes"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Activity Summary')
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Activity Summary']
        
        # Style the header row
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output.getvalue()


COS_API_KEY = "ehl6KMyT95fwzKf7sPW_X3eKFppy_24xbm4P1Yk-jqyU"
COS_SERVICE_INSTANCE_ID = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:projectreportnew"
COS_ENDPOINT = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
COS_BUCKET = "projectreportnew"


cos_client = ibm_boto3.client(
    's3',
    ibm_api_key_id=COS_API_KEY,
    ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
    config=Config(signature_version='oauth'),
    endpoint_url=COS_ENDPOINT
)

def get_cos_files():
    try:
        response = cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            print("No .json files found in the bucket 'ozonetell'. Please ensure JSON files are uploaded.")
        return files
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        return ["Error fetching COS files"]
    

if 'tower5df' not in st.session_state:
    st.session_state.tower5df = pd.DataFrame()
if 'tower4df' not in st.session_state:
    st.session_state.tower4df = pd.DataFrame()

foundveridiaf4 = False
foundveridiaf5 = False

today = date.today()
prev_month = today - relativedelta(months=1)

foundverdia = False

month_year = today.strftime("%m-%Y")
prev_month_year = prev_month.strftime("%m-%Y")


files = get_cos_files()
st.write(files)

tower4df = None
tower5df = None

if st.session_state.tower4df.empty:
    for file in files:
    
        try:
            if file.startswith("Veridia") and "Tower 5 Finishing Tracker" in file and month_year in file:
            
                st.write("✅ Current month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                st.session_state.tower5df, name = process_file_Tower5(io.BytesIO(response['Body'].read()))

                break
                # elif prev_month_year in file:
                #     st.write("🕓 Previous month:", file)
                #     response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                #     GetTower5Finishing(io.BytesIO(response['Body'].read()))
        except Exception as e:
            st.info(e)

    for file in files:
        
        try:
            if file.startswith("Veridia") and "Tower 4 Finishing Tracker" in file and prev_month_year in file:
            
                st.write("✅ Current month:", file)
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                st.session_state.tower4df, name = process_file_Tower4(io.BytesIO(response['Body'].read()))
                break
                # elif prev_month_year in file:
                #     st.write("🕓 Previous month:", file)
                #     response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                #     GetTower5Finishing(io.BytesIO(response['Body'].read()))
        except Exception as e:
            st.info(e)



# tower4json = tower4df.to_json(orient='records')
# tower5json = tower5df.to_json(orient='records')

# Combine available months and years from both dataframes
all_months = set()
all_years = set()

for key in ['tower5df', 'tower4df']:
    if key in st.session_state and not st.session_state[key].empty:
        df = st.session_state[key]
        all_months.update(df['Finish Month'].dropna().unique())
        all_years.update(df['Finish Year'].dropna().unique())

# Sort for display
all_months = sorted(all_months)
all_years = sorted(all_years)

# Sidebar filter UI
with st.sidebar:
    selected_months = st.multiselect("Select Finish Month(s)", all_months, default=all_months)
    selected_year = st.selectbox("Select Finish Year", all_years)

# Initialize empty DataFrames
filtered_df5 = pd.DataFrame()
filtered_df4 = pd.DataFrame()

# Filter and display tower5df
if 'tower5df' in st.session_state and not st.session_state.tower5df.empty:
    df5 = st.session_state.tower5df
    filtered_df5 = df5[
        (df5['Finish Month'].isin(selected_months)) &
        (df5['Finish Year'] == selected_year)
    ]
    st.subheader("Filtered tower5df")
    st.write(filtered_df5)

# Filter and display tower4df
if 'tower4df' in st.session_state and not st.session_state.tower4df.empty:
    df4 = st.session_state.tower4df
    filtered_df4 = df4[
        (df4['Finish Month'].isin(selected_months)) &
        (df4['Finish Year'] == selected_year)
    ]
    st.subheader("Filtered tower4df")
    st.write(filtered_df4)

# Combine and display as JSON
combined_df = pd.concat([filtered_df5, filtered_df4], ignore_index=True)

if not combined_df.empty:
    st.subheader("Combined Filtered Data (as JSON)")
    data = combined_df.to_dict(orient='records')
    activity_summary = process_activity_data(data)
    final_df = create_tower_format(activity_summary)
    
    # Display the formatted table
    st.subheader("Activity Summary by Month")
    st.dataframe(final_df, use_container_width=True)

    excel_data = to_excel(final_df)
    
    st.download_button(
        label="📥 Download as Excel",
        data=excel_data,
        file_name=f"activity_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("No data matching the selected filters.")
