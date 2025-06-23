import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from EWS_LIG import *
from Tower_G_and_H import *
from Veridia import *
from Wavecity import *
from Finishing import *
from Eden import *
from Fileformat import *
from datetime import date
import concurrent.futures
from dateutil.relativedelta import relativedelta
import re


if 'tower2_finishing' not in st.session_state:
    st.session_state.tower2_finishing = "0%"
if 'tower3_finishing' not in st.session_state:
    st.session_state.tower3_finishing = "0%"
if 'tower4_finishing' not in st.session_state:
    st.session_state.tower4_finishing = "0%"
if 'tower5_finishing' not in st.session_state:
    st.session_state.tower5_finishing = "0%"
if 'tower6_finishing' not in st.session_state:
    st.session_state.tower6_finishing = "0%"
if 'tower7_finishing' not in st.session_state:
    st.session_state.tower7_finishing = "0%"


if 'towerf_finishing' not in st.session_state:
    st.session_state.towerf_finishing = "0%"
if 'towerg_finishing' not in st.session_state:
    st.session_state.towerg_finishing = "0%"
if 'towerh_finishing' not in st.session_state:
    st.session_state.towerh_finishing = "0%"

if 'wavecity_finishing' not in st.session_state:
    st.session_state.wavecity_finishing = "0%"

COS_API_KEY = "ehl6KMyT95fwzKf7sPW_X3eKFppy_24xbm4P1Yk-jqyU"
COS_SERVICE_INSTANCE_ID = "crn:v1:bluemix:public:cloud-object-storage:global:a/fddc2a92db904306b413ed706665c2ff:e99c3906-0103-4257-bcba-e455e7ced9b7:bucket:projectreportnew"
COS_ENDPOINT = "https://s3.us-south.cloud-object-storage.appdomain.cloud"
COS_BUCKET = "projectreportnew"


def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Write the DataFrame to Excel, starting from row 1 to leave space for the title
        df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=1)
        
        # Get the xlsxwriter workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Define a format for the title row with yellow background
        title_format = workbook.add_format({
            'bold': True,
            'bg_color': 'yellow',
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # Add a title in the first row (e.g., "Tower Project Status")
        # worksheet.write(0, 0,0, f'Overall Project Report ({date.today()})', title_format)
        
        # Merge cells across the first row for the title (assuming the DataFrame has columns)
        worksheet.merge_range(0, 0, 0, len(df.columns)-1, f'Overall Project Report ({date.today()})', title_format)
        
    return output.getvalue()


cos_client = ibm_boto3.client(
    's3',
    ibm_api_key_id=COS_API_KEY,
    ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
    config=Config(signature_version='oauth'),
    endpoint_url=COS_ENDPOINT
)


def get_cos_files():
    try:
        response = cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            print("No .json files found in the bucket 'ozonetell'. Please ensure JSON files are uploaded.")
        return files
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        return ["Error fetching COS files",e]
    
files = get_cos_files()
# files = ["Error fetching COS files","Something Error"]
# files = ["EWS LIG P4/Structure Work Tracker (31-05-2025).xlsx", "Eden/Structure Work Tracker (31-05-2025).xlsx", "Eligo/Structure Work Tracker (31-05-2025).xlsx", "Eligo/Tower G Finishing Tracker (01-06-2025).xlsx", "Eligo/Tower H Finishing Tracker (01-06-2025).xlsx", "Veridia/Structure Work Tracker (31-05-2025).xlsx", "Veridia/Tower 4 Finishing Tracker (13-05-2025).xlsx", "Veridia/Tower 5 Finishing Tracker (01-06-2025).xlsx", "Veridia/Tower 7 Finishing Tracker (01-06-2025).xlsx", "Wave City Club/Structure Work Tracker Wave City Club all Block (11-06-2025).xlsx"]
# st.write(files)

today = datetime.today()
current_year = today.year
current_month = today.month
cutoff_day = 10



def GetOverallreport(files):
        ews_lig = {}
        veridia = {}
        eligo = {}
        Eden = {}
        wave = {}

        #VERIDIA TOWER 4
        for file in files:
            if file.startswith("Veridia") and "Tower 4 Finishing Tracker" in file:
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                GetTower4Finishing(io.BytesIO(response['Body'].read()))
                st.write(file,"✅")
                

        #VERIDIA TOWER 4
        for file in files:
            if file.startswith("Veridia") and "Tower 5 Finishing Tracker" in file:
               
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                GetTower5Finishing(io.BytesIO(response['Body'].read()))
                st.write(file,"✅")

        #VERIDIA TOWER 7
        for file in files:
            if file.startswith("Veridia") and "Tower 7 Finishing Tracker" in file:
               
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                GetTower7Finishing(io.BytesIO(response['Body'].read()))
                st.write(file,"✅")





        for file in files:
            if file.startswith("Eligo") and "Structure Work Tracker" in file:
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                eligo = ProcessGandH(io.BytesIO(response['Body'].read()))
                st.write(file,"✅")

    

        
        # #ELIGO TOWER G
        for file in files:
            if file.startswith("Eligo") and "Tower G Finishing Tracker" in file:
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                GetTowerGFinishing(io.BytesIO(response['Body'].read()))
                st.write(file,"✅")

        #ELIGO TOWER H
        for file in files:
            if file.startswith("Eligo") and "Tower H Finishing Tracker" in file:
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                GetTowerHFinishing(io.BytesIO(response['Body'].read()))
                st.write(file,"✅")


        for file in files:
            #WAVE CITY
            if file.startswith("Wave City Club") and "Structure Work Tracker Wave City Club all Block" in file:
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                wave = GetWaveCity(io.BytesIO(response['Body'].read()))
                # st.write(wave)
                st.write(file,"✅")

        # EWS LIG
        for file in files:
            if file.startswith("EWS LIG") and "Structure Work Tracker" in file:
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                ews_lig = ProcessEWSLIG(io.BytesIO(response['Body'].read()))
                st.write(file,"✅")

    

        # #ELIGO TOWER STRUCTURE
        # for file in files:
        #     if file.startswith("Eligo") and "Structure Work Tracker" in file:
        #         response = cos_client.get_object(Bucket="projectreportnew", Key=file)
        #         eligo = ProcessGandH(io.BytesIO(response['Body'].read()))
        #         st.write(file,"✅")


        #EDEN
        for file in files:
             if file.startswith("Eden") and "Structure Work Tracker" in file:
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                Eden = get_percentages(io.BytesIO(response['Body'].read()))


       
        # #VERIDIA TOWER 4
        for file in files:
            if file.startswith("Veridia") and "Structure Work Tracker" in file:
                response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                veridia = ProcessVeridia(io.BytesIO(response['Body'].read()))
                # st.write(veridia)
                st.write(file,"✅")

        # for i in files_after_or_on_10th:
        #     st.write(i)

        combined_data = []

        for data in [ews_lig, veridia, eligo, Eden, wave]:
            if isinstance(data, list):
                # Check if it's a list of dicts (most typical for DataFrame input)
                if all(isinstance(item, dict) for item in data):
                    combined_data.extend(data)
                else:
                    st.warning("Some Files Are Missing")
            elif data is not None:
                st.warning("Some Files Are Missing")

        # Now safely create the DataFrame
        if combined_data:
            df = pd.DataFrame(combined_data)
            return df


st.header("OVERALL PROJECT REPORT")
def extract_date(filename):
    # Regex to find date in format dd-mm-yyyy or dd-mm-yyyy in parentheses or after dash
    match = re.search(r'(\d{2})-(\d{2})-(\d{4})', filename)
    if match:
        day, month, year = map(int, match.groups())
        return datetime(year, month, day)
    return None

files_before_10th = []
files_after_or_on_10th = []

if files[0] == "Error fetching COS files":
    st.write("")
else:
    for f in files:
        file_date = extract_date(f)
        if file_date:
            # Filter for current month and year only
            if file_date.year == current_year and file_date.month == current_month:
                if file_date.day < cutoff_day:
                    files_before_10th.append(f)
                else:
                    files_after_or_on_10th.append(f)
            else:
                # For files not in current month/year, treat as before 10th
                files_before_10th.append(f)
        else:
            # If no date found, skip or decide where to put
            pass

# st.write("Files BEFORE 10th June 2025:")
# for file in files_before_10th:
#     st.write(file)

# st.write("\nFiles ON or AFTER 10th June 2025:")
# for file in files_after_or_on_10th:
#     st.write(file)


st.info(f"Previous Month Files:{len(files_before_10th)}")
st.success(f"Current Month Files:{len(files_after_or_on_10th)}")

veridia_finishing_4 = []
veridia_finishing_5 = []
eligo_finishing_g = []
eligo_finishing_h = []
wave_city = []
ews_lig_structure = []
eligo_structure = []
eden_structure = []
veridia_structure = []

if 'overall' not in st.session_state:
    st.session_state.overall = None

if 'overalldf' not in st.session_state:
    st.session_state.overalldf = None

if 'check' not in st.session_state:
    st.session_state.check = False

def seperatefiles(files):
        
        #VERIDIA TOWER 4
        for file in files_before_10th:
            if file.startswith("Veridia") and "Tower 4 Finishing Tracker" in file:
                veridia_finishing_4.append(file)

        #VERIDIA TOWER 4
        for file in files_before_10th:
            if file.startswith("Veridia") and "Tower 5 Finishing Tracker" in file:
                veridia_finishing_5.append(file)

        
        #ELIGO TOWER G
        for file in files_before_10th:
            if file.startswith("Eligo") and "Tower G Finishing Tracker" in file:
                eligo_finishing_g.append(file)

        #ELIGO TOWER H
        for file in files_before_10th:
            if file.startswith("Eligo") and "Tower H Finishing Tracker" in file:
                eligo_finishing_h.append(file)

        for file in files_before_10th:
            #WAVE CITY
            if file.startswith("Wave City Club") and "Structure Work Tracker Wave City Club all Block" in file:
                wave_city.append(file)

        # EWS LIG
        for file in files_before_10th:
            if file.startswith("EWS LIG") and "Structure Work Tracker" in file:
                ews_lig_structure.append(file)

    

        #ELIGO TOWER STRUCTURE
        for file in files_before_10th:
            if file.startswith("Eligo") and "Structure Work Tracker" in file:
                eligo_structure.append(file)

        #EDEN
        for file in files_before_10th:
            if file.startswith("Eden") and "Structure Work Tracker" in file:
                eden_structure.append(file)


       
        # #VERIDIA TOWER 4
        for file in files_before_10th:
            if file.startswith("Veridia") and "Structure Work Tracker" in file:
                veridia_structure.append(file)


if files[0] == "Error fetching COS files":
    st.warning(files[1])     
else:
    if len(files_before_10th) > 0:
        seperatefiles(files)
        if len(veridia_finishing_4) > 1 or len(eligo_finishing_g) > 1 or len(eligo_finishing_h) > 1 or len(wave_city) > 1 or len(ews_lig_structure) > 1 or len(eligo_structure) > 1 or len(eden_structure) > 1 or len(veridia_structure) > 1:
            st.warning("There are multiple files for the same project. Please check the files and remove duplicates")
        # st.write(files_before_10th)
            with st.form("my_form"):
                st.info("Multiple Previous Files Found Please Select a files to continue")
                selected_files = st.multiselect("Choose a Files",files)
                # Every form must have a submit button.
                submitted = st.form_submit_button("Continue",type="primary",use_container_width=True)
                if submitted:
                    # st.write(selected_files)
                    st.session_state.overalldf = GetOverallreport(selected_files)
                    st.session_state.check = True
                    
            if st.session_state.check:
                if st.session_state.overalldf is not None and not st.session_state.overalldf.empty:
                # st.write(df)
                    excel_data = to_excel(st.session_state.overalldf)
                    st.session_state.overall = excel_data

                    st.title("Tower Project Status Table")

                    st.dataframe(st.session_state.overalldf)

                    st.download_button(
                        label="Download as Excel",
                        data=excel_data,
                        file_name="Overall_Project_Report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    else:
        st.session_state.overalldf = GetOverallreport(files)
        st.session_state.check = True
        if st.session_state.check:
            if st.session_state.overalldf is not None and not st.session_state.overalldf.empty:
                st.dataframe(st.session_state.overalldf)
            # st.write(df)
                excel_data = to_excel(st.session_state.overalldf)
                st.session_state.overall = excel_data
                st.title("Tower Project Status Table")

                # st.dataframe(df)

                st.download_button(
                    label="Download as Excel",
                    data=excel_data,
                    file_name="Overall_Project_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                   

      

        